import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime
import hashlib
import sqlite3
import os
from pathlib import Path
import tempfile
import shutil
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

# Page configuration
st.set_page_config(
    page_title="Running Analysis Suite",
    page_icon="🏃",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Prompt:wght@300;400;500;600;700&display=swap');

    * {
        font-family: 'Prompt', sans-serif !important;
    }

    .stApp {
        background-color: rgb(255, 242, 224);
    }

    .stSidebar {
        background-color: rgb(192, 201, 238);
    }

    .stButton > button {
        background-color: rgb(137, 138, 196);
        color: white;
        border: none;
        border-radius: 5px;
        transition: all 0.3s;
    }

    .stButton > button:hover {
        background-color: rgb(162, 170, 219);
        transform: translateY(-2px);
    }

    .metric-card {
        background-color: white;
        padding: 20px;
        border-radius: 10px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        margin: 10px 0;
    }

    .user-badge {
        background-color: rgb(137, 138, 196);
        color: white;
        padding: 5px 15px;
        border-radius: 20px;
        display: inline-block;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'user_type' not in st.session_state:
    st.session_state.user_type = None
if 'username' not in st.session_state:
    st.session_state.username = None


# Database setup
def init_db():
    conn = sqlite3.connect('running_analysis.db')
    c = conn.cursor()

    # Users table
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (
                     id
                     INTEGER
                     PRIMARY
                     KEY
                     AUTOINCREMENT,
                     username
                     TEXT
                     UNIQUE
                     NOT
                     NULL,
                     password
                     TEXT
                     NOT
                     NULL,
                     user_type
                     TEXT
                     NOT
                     NULL,
                     created_at
                     TIMESTAMP
                     DEFAULT
                     CURRENT_TIMESTAMP
                 )''')

    # Runners table
    c.execute('''CREATE TABLE IF NOT EXISTS runners
    (
        id
        INTEGER
        PRIMARY
        KEY
        AUTOINCREMENT,
        name
        TEXT
        NOT
        NULL,
        coach_id
        INTEGER,
        created_at
        TIMESTAMP
        DEFAULT
        CURRENT_TIMESTAMP,
        FOREIGN
        KEY
                 (
        coach_id
                 ) REFERENCES users
                 (
                     id
                 ))''')

    # Performance data table
    c.execute('''CREATE TABLE IF NOT EXISTS performance_data
    (
        id
        INTEGER
        PRIMARY
        KEY
        AUTOINCREMENT,
        runner_id
        INTEGER,
        test_date
        TIMESTAMP,
        range_0_25
        TEXT,
        range_25_50
        TEXT,
        range_50_75
        TEXT,
        range_75_100
        TEXT,
        max_speed
        REAL,
        avg_speed
        REAL,
        total_time
        REAL,
        created_at
        TIMESTAMP
        DEFAULT
        CURRENT_TIMESTAMP,
        FOREIGN
        KEY
                 (
        runner_id
                 ) REFERENCES runners
                 (
                     id
                 ))''')

    # Insert default admin user
    try:
        c.execute("INSERT INTO users (username, password, user_type) VALUES (?, ?, ?)",
                  ('admin', hashlib.sha256('admin123'.encode()).hexdigest(), 'admin'))
    except:
        pass

    conn.commit()
    conn.close()


# Initialize database
init_db()


# Authentication functions
def authenticate_user(username, password):
    conn = sqlite3.connect('running_analysis.db')
    c = conn.cursor()
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    c.execute("SELECT id, user_type FROM users WHERE username = ? AND password = ?",
              (username, hashed_password))
    result = c.fetchone()
    conn.close()
    return result


def register_user(username, password, user_type):
    conn = sqlite3.connect('running_analysis.db')
    c = conn.cursor()
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    try:
        c.execute("INSERT INTO users (username, password, user_type) VALUES (?, ?, ?)",
                  (username, hashed_password, user_type))
        conn.commit()
        conn.close()
        return True
    except:
        conn.close()
        return False


# Parse performance data from text file
def parse_performance_data(content):
    lines = content.strip().split('\n')
    data = []

    for line in lines[1:]:  # Skip header
        parts = line.split()
        if len(parts) >= 4:
            try:
                data.append({
                    'time': float(parts[0]),
                    'mass_A': float(parts[1]),
                    't': float(parts[2]),
                    'x': float(parts[3]),
                    'v': float(parts[4]) if len(parts) > 4 else 0
                })
            except:
                continue

    return pd.DataFrame(data)


# Generate performance visualization
def create_performance_chart(data_dict):
    fig = go.Figure()

    colors = ['#5A7ECF', '#E68A5C', '#A8A8A8', '#F0D050']
    ranges = ['0-25m', '25-50m', '50-75m', '75-100m']

    for i, (range_name, df) in enumerate(data_dict.items()):
        if df is not None and not df.empty:
            fig.add_trace(go.Scatter(
                x=df['time'],
                y=df['v'],
                mode='lines+markers',
                name=ranges[i],
                line=dict(color=colors[i], width=3),
                marker=dict(size=8)
            ))

    fig.update_layout(
        title='100-Meter Speed Test',
        xaxis_title='Time (s)',
        yaxis_title='Speed (m/s)',
        height=500,
        plot_bgcolor='white',
        paper_bgcolor='rgb(255, 242, 224)',
        font=dict(family='Prompt', size=12),
        showlegend=True,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        )
    )

    fig.update_xaxis(showgrid=True, gridwidth=1, gridcolor='lightgray')
    fig.update_yaxis(showgrid=True, gridwidth=1, gridcolor='lightgray')

    return fig


# Generate Excel report
def generate_excel_report(data_dict, runner_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Report"

    # Header styling
    header_font = Font(name='Prompt', size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8A8AC4", end_color="8A8AC4", fill_type="solid")

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Running Performance Report - {runner_name}"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Date
    ws['A2'] = f"Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Performance summary
    row = 4
    ws[f'A{row}'] = "Performance Summary"
    ws[f'A{row}'].font = Font(name='Prompt', size=14, bold=True)

    row += 2
    headers = ['Range', 'Max Speed (m/s)', 'Avg Speed (m/s)', 'Time (s)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name='Prompt', bold=True)
        cell.fill = PatternFill(start_color="C0C9EE", end_color="C0C9EE", fill_type="solid")

    # Data rows
    row += 1
    ranges = ['0-25m', '25-50m', '50-75m', '75-100m']
    for i, (range_name, df) in enumerate(data_dict.items()):
        if df is not None and not df.empty:
            ws.cell(row=row + i, column=1, value=ranges[i])
            ws.cell(row=row + i, column=2, value=round(df['v'].max(), 2))
            ws.cell(row=row + i, column=3, value=round(df['v'].mean(), 2))
            ws.cell(row=row + i, column=4, value=round(df['time'].max(), 2))

    # Detailed data
    row += 6
    ws[f'A{row}'] = "Detailed Performance Data"
    ws[f'A{row}'].font = Font(name='Prompt', size=14, bold=True)

    row += 2
    col = 1
    for range_name, df in data_dict.items():
        if df is not None and not df.empty:
            # Range header
            ws.cell(row=row, column=col, value=range_name)
            ws.cell(row=row, column=col).font = Font(name='Prompt', bold=True)

            # Column headers
            headers = ['Time', 'Speed']
            for j, header in enumerate(headers):
                ws.cell(row=row + 1, column=col + j, value=header)
                ws.cell(row=row + 1, column=col + j).font = Font(name='Prompt', bold=True)

            # Data
            for idx, data_row in df.iterrows():
                ws.cell(row=row + 2 + idx, column=col, value=round(data_row['time'], 3))
                ws.cell(row=row + 2 + idx, column=col + 1, value=round(data_row['v'], 2))

            col += 3

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


# Login page
def login_page():
    st.title("🏃 Running Analysis Suite")
    st.markdown("<h3 style='text-align: center;'>Login to Continue</h3>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        with st.form("login_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login", use_container_width=True)

            if submitted:
                result = authenticate_user(username, password)
                if result:
                    st.session_state.authenticated = True
                    st.session_state.user_id = result[0]
                    st.session_state.user_type = result[1]
                    st.session_state.username = username
                    st.rerun()
                else:
                    st.error("Invalid username or password")

        st.markdown("---")
        st.markdown("Don't have an account? Contact admin for registration.")


# Main application
def main_app():
    # Sidebar
    with st.sidebar:
        st.markdown(f"<div class='user-badge'>{st.session_state.user_type.upper()}: {st.session_state.username}</div>",
                    unsafe_allow_html=True)

        if st.button("Logout", use_container_width=True):
            st.session_state.authenticated = False
            st.session_state.user_type = None
            st.session_state.username = None
            st.rerun()

        st.markdown("---")

        # Navigation
        if st.session_state.user_type == 'admin':
            page = st.selectbox("Navigation",
                                ["Upload & Analyze", "View Reports", "Manage Users", "Manage Runners"])
        elif st.session_state.user_type == 'coach':
            page = st.selectbox("Navigation",
                                ["Upload & Analyze", "View Reports", "My Runners"])
        else:
            page = st.selectbox("Navigation",
                                ["Upload & Analyze", "View Reports"])

    # Main content
    if page == "Upload & Analyze":
        upload_analyze_page()
    elif page == "View Reports":
        view_reports_page()
    elif page == "Manage Users" and st.session_state.user_type == 'admin':
        manage_users_page()
    elif page == "Manage Runners" and st.session_state.user_type == 'admin':
        manage_runners_page()
    elif page == "My Runners" and st.session_state.user_type == 'coach':
        my_runners_page()


# Upload and analyze page
def upload_analyze_page():
    st.title("📹 Upload & Analyze Performance")

    # Runner selection
    conn = sqlite3.connect('running_analysis.db')
    c = conn.cursor()

    if st.session_state.user_type == 'coach':
        c.execute("""SELECT r.id, r.name
                     FROM runners r
                              JOIN users u ON r.coach_id = u.id
                     WHERE u.username = ?""", (st.session_state.username,))
    else:
        c.execute("SELECT id, name FROM runners")

    runners = c.fetchall()
    conn.close()

    if not runners:
        st.warning("No runners found. Please add runners first.")
        return

    runner_dict = {name: id for id, name in runners}
    selected_runner = st.selectbox("Select Runner", list(runner_dict.keys()))

    st.markdown("---")

    # File upload section
    st.subheader("Upload Video Files")
    st.info("Please upload 4 video files for each range of the 100m sprint")

    col1, col2 = st.columns(2)

    with col1:
        video_0_25 = st.file_uploader("0-25m Video", type=['mp4', 'avi', 'mov'])
        video_25_50 = st.file_uploader("25-50m Video", type=['mp4', 'avi', 'mov'])

    with col2:
        video_50_75 = st.file_uploader("50-75m Video", type=['mp4', 'avi', 'mov'])
        video_75_100 = st.file_uploader("75-100m Video", type=['mp4', 'avi', 'mov'])

    # For demo purposes, also allow text file upload
    st.markdown("---")
    st.subheader("Demo: Upload Performance Data (Text Files)")

    col1, col2 = st.columns(2)

    with col1:
        data_0_25 = st.file_uploader("0-25m Data", type=['txt'])
        data_25_50 = st.file_uploader("25-50m Data", type=['txt'])

    with col2:
        data_50_75 = st.file_uploader("50-75m Data", type=['txt'])
        data_75_100 = st.file_uploader("75-100m Data", type=['txt'])

    if st.button("Analyze Performance", use_container_width=True):
        with st.spinner("Processing data..."):
            # Process uploaded data
            data_dict = {}

            # Parse text files if uploaded
            if data_0_25:
                content = data_0_25.read().decode('utf-8')
                data_dict['0-25'] = parse_performance_data(content)
            else:
                # Generate sample data
                time_range = np.linspace(0, 3, 50)
                speeds = 7.06 + np.random.normal(0, 0.2, 50)
                data_dict['0-25'] = pd.DataFrame({'time': time_range, 'v': speeds})

            if data_25_50:
                content = data_25_50.read().decode('utf-8')
                data_dict['25-50'] = parse_performance_data(content)
            else:
                time_range = np.linspace(3, 6, 50)
                speeds = 8.57 + np.random.normal(0, 0.2, 50)
                data_dict['25-50'] = pd.DataFrame({'time': time_range, 'v': speeds})

            if data_50_75:
                content = data_50_75.read().decode('utf-8')
                data_dict['50-75'] = parse_performance_data(content)
            else:
                time_range = np.linspace(6, 9, 50)
                speeds = 8.56 + np.random.normal(0, 0.2, 50)
                data_dict['50-75'] = pd.DataFrame({'time': time_range, 'v': speeds})

            if data_75_100:
                content = data_75_100.read().decode('utf-8')
                data_dict['75-100'] = parse_performance_data(content)
            else:
                time_range = np.linspace(9, 12, 50)
                speeds = 8.11 + np.random.normal(0, 0.3, 50)
                data_dict['75-100'] = pd.DataFrame({'time': time_range, 'v': speeds})

            # Display results
            st.success("Analysis Complete!")

            # Performance metrics
            st.subheader("Performance Metrics")
            col1, col2, col3, col4 = st.columns(4)

            metrics = []
            for i, (range_name, df) in enumerate(data_dict.items()):
                if df is not None and not df.empty:
                    max_speed = df['v'].max()
                    avg_speed = df['v'].mean()
                    metrics.append((range_name, max_speed, avg_speed))

            for i, (range_name, max_speed, avg_speed) in enumerate(metrics):
                with [col1, col2, col3, col4][i]:
                    st.metric(f"{range_name}m", f"{max_speed:.2f} m/s", f"Avg: {avg_speed:.2f}")

            # Visualization
            st.subheader("Speed Analysis")
            fig = create_performance_chart(data_dict)
            st.plotly_chart(fig, use_container_width=True)

            # Generate report
            excel_report = generate_excel_report(data_dict, selected_runner)

            st.download_button(
                label="📊 Download Excel Report",
                data=excel_report,
                file_name=f"performance_report_{selected_runner}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Save to database
            conn = sqlite3.connect('running_analysis.db')
            c = conn.cursor()

            # Calculate overall metrics
            all_speeds = []
            for df in data_dict.values():
                if df is not None and not df.empty:
                    all_speeds.extend(df['v'].tolist())

            max_speed = max(all_speeds) if all_speeds else 0
            avg_speed = sum(all_speeds) / len(all_speeds) if all_speeds else 0
            total_time = 12.0  # Approximate for demo

            c.execute("""INSERT INTO performance_data
                             (runner_id, test_date, max_speed, avg_speed, total_time)
                         VALUES (?, ?, ?, ?, ?)""",
                      (runner_dict[selected_runner], datetime.now(), max_speed, avg_speed, total_time))

            conn.commit()
            conn.close()

            st.success("Data saved to database!")


# View reports page
def view_reports_page():
    st.title("📊 Performance Reports")

    conn = sqlite3.connect('running_analysis.db')

    # Get performance data based on user type
    if st.session_state.user_type == 'coach':
        query = """
                SELECT p.*, r.name as runner_name
                FROM performance_data p
                         JOIN runners r ON p.runner_id = r.id
                         JOIN users u ON r.coach_id = u.id
                WHERE u.username = ?
                ORDER BY p.test_date DESC \
                """
        df = pd.read_sql_query(query, conn, params=(st.session_state.username,))
    else:
        query = """
                SELECT p.*, r.name as runner_name
                FROM performance_data p
                         JOIN runners r ON p.runner_id = r.id
                ORDER BY p.test_date DESC \
                """
        df = pd.read_sql_query(query, conn)

    conn.close()

    if df.empty:
        st.info("No performance data available yet.")
        return

    # Filter options
    col1, col2 = st.columns(2)
    with col1:
        runners = df['runner_name'].unique()
        selected_runner = st.selectbox("Filter by Runner", ["All"] + list(runners))

    with col2:
        date_range = st.date_input("Date Range",
                                   value=(df['test_date'].min(), df['test_date'].max()),
                                   format="YYYY-MM-DD")

    # Apply filters
    if selected_runner != "All":
        df = df[df['runner_name'] == selected_runner]

    # Display summary statistics
    st.subheader("Summary Statistics")
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Total Tests", len(df))
    with col2:
        st.metric("Avg Max Speed", f"{df['max_speed'].mean():.2f} m/s")
    with col3:
        st.metric("Best Speed", f"{df['max_speed'].max():.2f} m/s")
    with col4:
        st.metric("Avg Time", f"{df['total_time'].mean():.2f} s")

    # Performance trend
    st.subheader("Performance Trends")

    if selected_runner != "All" and len(df) > 1:
        fig = go.Figure()

        fig.add_trace(go.Scatter(
            x=pd.to_datetime(df['test_date']),
            y=df['max_speed'],
            mode='lines+markers',
            name='Max Speed',
            line=dict(color='rgb(137, 138, 196)', width=3)
        ))

        fig.add_trace(go.Scatter(
            x=pd.to_datetime(df['test_date']),
            y=df['avg_speed'],
            mode='lines+markers',
            name='Avg Speed',
            line=dict(color='rgb(192, 201, 238)', width=3)
        ))

        fig.update_layout(
            title=f"Speed Progression - {selected_runner}",
            xaxis_title="Date",
            yaxis_title="Speed (m/s)",
            height=400,
            plot_bgcolor='white',
            paper_bgcolor='rgb(255, 242, 224)'
        )

        st.plotly_chart(fig, use_container_width=True)

    # Detailed table
    st.subheader("Detailed Records")
    display_df = df[['runner_name', 'test_date', 'max_speed', 'avg_speed', 'total_time']]
    display_df.columns = ['Runner', 'Test Date', 'Max Speed (m/s)', 'Avg Speed (m/s)', 'Total Time (s)']
    st.dataframe(display_df, use_container_width=True)


# Manage users page (admin only)
def manage_users_page():
    st.title("👥 Manage Users")

    tab1, tab2 = st.tabs(["View Users", "Add User"])

    with tab1:
        conn = sqlite3.connect('running_analysis.db')
        users_df = pd.read_sql_query("SELECT id, username, user_type, created_at FROM users", conn)
        conn.close()

        st.dataframe(users_df, use_container_width=True)

    with tab2:
        st.subheader("Add New User")

        with st.form("add_user_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            user_type = st.selectbox("User Type", ["runner", "coach", "admin"])

            if st.form_submit_button("Add User"):
                if register_user(username, password, user_type):
                    st.success(f"User '{username}' added successfully!")
                    st.rerun()
                else:
                    st.error("Failed to add user. Username might already exist.")


# Manage runners page (admin only)
def manage_runners_page():
    st.title("🏃 Manage Runners")

    tab1, tab2 = st.tabs(["View Runners", "Add Runner"])

    with tab1:
        conn = sqlite3.connect('running_analysis.db')
        runners_df = pd.read_sql_query("""
                                       SELECT r.id, r.name, u.username as coach, r.created_at
                                       FROM runners r
                                                LEFT JOIN users u ON r.coach_id = u.id
                                       """, conn)
        conn.close()

        st.dataframe(runners_df, use_container_width=True)

    with tab2:
        st.subheader("Add New Runner")

        conn = sqlite3.connect('running_analysis.db')
        c = conn.cursor()
        c.execute("SELECT id, username FROM users WHERE user_type = 'coach'")
        coaches = c.fetchall()
        conn.close()

        with st.form("add_runner_form"):
            runner_name = st.text_input("Runner Name")

            if coaches:
                coach_dict = {username: id for id, username in coaches}
                selected_coach = st.selectbox("Assign to Coach", list(coach_dict.keys()))
            else:
                st.warning("No coaches available. Please add a coach first.")
                selected_coach = None

            if st.form_submit_button("Add Runner"):
                if runner_name and selected_coach:
                    conn = sqlite3.connect('running_analysis.db')
                    c = conn.cursor()
                    c.execute("INSERT INTO runners (name, coach_id) VALUES (?, ?)",
                              (runner_name, coach_dict[selected_coach]))
                    conn.commit()
                    conn.close()
                    st.success(f"Runner '{runner_name}' added successfully!")
                    st.rerun()
                else:
                    st.error("Please fill all fields.")


# My runners page (coach only)
def my_runners_page():
    st.title("👥 My Runners")

    conn = sqlite3.connect('running_analysis.db')

    # Get coach's runners
    runners_df = pd.read_sql_query("""
                                   SELECT r.id,
                                          r.name,
                                          COUNT(p.id)      as total_tests,
                                          MAX(p.max_speed) as best_speed,
                                          AVG(p.avg_speed) as avg_speed
                                   FROM runners r
                                            JOIN users u ON r.coach_id = u.id
                                            LEFT JOIN performance_data p ON r.id = p.runner_id
                                   WHERE u.username = ?
                                   GROUP BY r.id, r.name
                                   """, conn, params=(st.session_state.username,))

    conn.close()

    if runners_df.empty:
        st.info("No runners assigned to you yet.")
        return

    # Display runners with their stats
    for _, runner in runners_df.iterrows():
        with st.expander(f"🏃 {runner['name']}"):
            col1, col2, col3 = st.columns(3)

            with col1:
                st.metric("Total Tests", int(runner['total_tests']))
            with col2:
                best_speed = runner['best_speed'] if runner['best_speed'] else 0
                st.metric("Best Speed", f"{best_speed:.2f} m/s")
            with col3:
                avg_speed = runner['avg_speed'] if runner['avg_speed'] else 0
                st.metric("Avg Speed", f"{avg_speed:.2f} m/s")


# Main execution
if __name__ == "__main__":
    if st.session_state.authenticated:
        main_app()
    else:
        login_page()